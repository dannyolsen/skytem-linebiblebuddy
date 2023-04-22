import pandas as pd
import numpy as np
from pprint import pprint
import os
from openpyxl import load_workbook
import shutil
import tkinter as tk
from tkinter import filedialog

#FUNCTIONS
def extract_linedata_production(path_to_txt):
    all_files       = os.listdir(path_to_txt)
    txt_files       = [file for file in all_files if file.endswith('.txt')]

    # get the absolute path of the input directory
    #txt_dir = os.path.abspath(input_directory)
    txt_dir = os.path.dirname(path_to_txt)

    # open the output file in write mode
    with open('linedata_production.txt', 'w') as output_file:
        # iterate over each file name in the list
        header_written  = False
        write = False
        for file_name in txt_files:
            print(os.path.join(path_to_txt,file_name))
            # open the file in read mode and read its contents
            with open(os.path.join(path_to_txt,file_name), 'r') as f:
                # print(file_name)
                lines = f.readlines()

            if header_written == False:
                for line in lines:
                    if line.startswith('PRODUCTION LINES'):
                        write = True
                        continue

                    if write:
                        if line.startswith('        '):  # 8 spaces
                            write = False

                        else:
                            output_file.write(line)

                header_written = True

            elif header_written == True:
                lines_iter = iter(lines) # create an iterator from the list
                for line in lines_iter:
                    if line.startswith('PRODUCTION LINES'):
                        # skip two lines
                        for _ in range(1):
                            next(lines_iter, None)
                        write = True
                        continue

                    if write:
                        if line.startswith('        '):  # 8 spaces
                            write = False

                        else:
                            output_file.write(line)
                            # print(line)

def read_txt(production_txt, hight_line_range, freestyle_line_range):
    #READ TEXTFILE
    # read the text file data into a pandas DataFrame
    df_textfile                             = pd.read_csv(production_txt, sep='\s+')
    df_textfile                             = df_textfile.sort_values('Mission', ascending=True)

    # Specify the range of line numbers
    # freestyle lines
    # 400 meter lines
    # 1000 m lines
    # hight_line_range = [(920000, 929999),
    #               (930000, 939999)]
    
    hight_line_range_masked = hight_line_range.copy()
    
    # copy freestyle line values to 'mission' so program will insert this into excel when that time comes
    for index, row in df_textfile.iterrows():
        for range_item in freestyle_line_range:
            if row['Line'] >= range_item[0] and row['Line'] <= range_item[1]:
                df_textfile.loc[index, 'Mission'] = row['Line']
                break

    # Loop through the line ranges and set the 'Mission' column to None - set to None 400 m and 1000 m
    for lr in hight_line_range:
        line_mask = (df_textfile["Line"] >= lr[0]) & (df_textfile["Line"] <= lr[1])
        df_textfile.loc[line_mask, 'Mission'] = '*'
        hight_line_range_masked.remove(lr)

    df_textfile['Occurrence'] = df_textfile.groupby('Mission').cumcount() + 1
    df_textfile.to_csv('df_textfile.csv', sep=' ', index=False)

    return df_textfile

def get_amount_of_block_sheets(workbook):
    sheetcounter = 0
    for i in range(100):
        sheet_name          = 'Block'+str(i+1)
        # Check if a sheet with a specific name exists in the workbook
        if sheet_name in workbook.sheetnames:
            #print(sheet_name + ' exists in the workbook')
            sheetcounter += 1
            #print(sheetcounter)
        else:
            print("There are {} sheets in the workbook".format(sheetcounter))
            break
    
    return sheetcounter

def backup_linebible(original_file_path):
    #cwd             = os.path.dirname(original_file_path)

    # create a backup file path with the word "backup" added
    backup_file_path = original_file_path[:-5] + "_backup.xlsx"

    # create a copy of the original file with the new file name
    shutil.copy(original_file_path, backup_file_path)

def choose_linebible():
    # Prompt the user to select a directory
    txt_path = filedialog.askdirectory(title="Select dir with your txt data files")

    # Prompt the user to select an Excel file
    file_path = filedialog.askopenfilename(initialdir=txt_path, title="Select Excel File", filetypes=[("Excel files", "*.xlsx")])

    return file_path, txt_path

def get_missing_excel_number():
    #HER!!!!!!!
    pass

def get_all_mission_numbers(sheet_range):
    #create a list of all mission numbers to see if we have mission lines that does not exist in excel sheet
    all_mission_numbers = []
    for i in range(sheet_range):
        sheet_name          = 'Block'+str(i+1)
        header_row_index    = 2
        df_linebible        = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row_index)
        
        # strip leading and trailing spaces from string columns
        df_linebible        = df_linebible.applymap(lambda x: x.strip() if type(x) == str else x)
        for i, number in enumerate(df_linebible['Number']):
            all_mission_numbers.append(number)
        
    return all_mission_numbers

def main():
    # MAIN

    from art import text2art
    ##################################################################################################
    print(text2art('LineBibleBuddy'))
    print('-------------------------------------------------------------')
    print('22 April 2023 - Danny Olsen (dol@skytem.com)')
    print('-------------------------------------------------------------\n')
    print('This program will edit your linebible and insert datavalues.\n')
    ##################################################################################################

    #Change to selected linebible dir
    os.chdir(cwd)

    #create a txt file that we can work with from one or multiple txt files
    extract_linedata_production(path_to_txt)

    #put all data from txt into dataframe
    df_textfile                 = read_txt("linedata_production.txt", hight_line_range, freestyle_line_range)

    print("Checking if all values exist in excel sheet")

    sheet_range                 = get_amount_of_block_sheets(workbook)

    df_all_mission_numbers      = pd.DataFrame(get_all_mission_numbers(sheet_range), columns=['mission_numbers'])

    #df_all_mission_numbers      = pd.DataFrame(all_mission_numbers, columns=['mission_numbers'])
    df_textfile['Mission']      = df_textfile['Mission'].astype(str)

    missing_missions            = set(df_textfile['Mission']) - set(df_all_mission_numbers['mission_numbers'])
    
    for mis in missing_missions:
        if mis.isdigit():       #NOT WORKING CORRECTLY WILL NOT CATCH FREESTYLE LINES NOT IN EXCEL
            myInt = int(mis)
            in_range = False
            for start, end in hight_line_range + freestyle_line_range:
                if start <= myInt <= end:
                    in_range = True
                    break
            if not in_range:
                print(myInt)

        elif mis == '*':
            pass

        elif isinstance(mis, str):
            print(str(mis) + " not in linebible")

        else:
            print(type(mis))
            print(str(mis) + " not in linebible")

    for i in range(sheet_range):
        try:
            sheet_name          = 'Block'+str(i+1)
            header_row_index    = 2
            df_linebible        = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row_index)

            #READ EXCEL AND CREATING PROPER DATAFRAME TO WORK WITH######
            new_names = {}
            counter = 0
            for name in df_linebible.columns:
                #if name.startswith('km'):
                if name.startswith('km'):
                    counter += 1
                    new_names[name] = f'km.{counter}'

            df_linebible        = df_linebible.rename(columns=new_names)

            # strip leading and trailing spaces from string columns
            df_linebible        = df_linebible.applymap(lambda x: x.strip() if type(x) == str else x)

            # sort df_textfile by Mission column
            df_textfile['Mission'] = df_textfile['Mission'].astype(str)
            df_textfile = df_textfile.sort_values('Mission', ascending=True)

            for index, row in df_textfile.iterrows():
                mission         = row['Mission']
                flown_heli      = row['FlownHel']
                occurrence      = row['Occurrence']
                flight          = row['Flight']
                #line            = row['Line']

                for i, number in enumerate(df_linebible['Number']):
                    if mission == number:
                        #finder det row index der passer
                        row_index = df_linebible.loc[df_linebible['Number'] == mission].index[0]

                        flt_col = 'FLT' + str(int(occurrence))
                        km_col = 'km.' + str(int(occurrence))

                        #skrives til df_linebible på det rigtige row index
                        df_linebible.at[row_index, flt_col] = flight
                        df_linebible.at[row_index, km_col] = flown_heli
                        #print(i,occurrence, mission, flt_col, df_linebible.loc[i, flt_col], km_col, df_linebible.loc[i, km_col])

            #df_linebible.to_csv('df_linebible.csv', sep=' ', index=False)

            # Select worksheet to insert data - this takes a long time
            print("Writing to workbook sheet : " + sheet_name)
            worksheet = workbook[sheet_name]
            col = ['E','H','K','N','Q','T']
            # Insert dataframe into Excel worksheet

            for i, name in enumerate(df_linebible['FLT1'], start=4):
                worksheet[f'{col[0]}{i}'] = name

            for i, name in enumerate(df_linebible['FLT2'], start=4):
                worksheet[f'{col[1]}{i}'] = name

            for i, name in enumerate(df_linebible['FLT3'], start=4):
                worksheet[f'{col[2]}{i}'] = name

            for i, name in enumerate(df_linebible['FLT4'], start=4):
                worksheet[f'{col[3]}{i}'] = name

            for i, name in enumerate(df_linebible['FLT5'], start=4):
                worksheet[f'{col[4]}{i}'] = name

            for i, name in enumerate(df_linebible['FLT6'], start=4):
                worksheet[f'{col[5]}{i}'] = name

            col = ['F','I','L','O','R','U']
            # Insert dataframe into Excel worksheet

            for i, name in enumerate(df_linebible['km.1'], start=4):
                worksheet[f'{col[0]}{i}'] = name

            for i, name in enumerate(df_linebible['km.2'], start=4):
                worksheet[f'{col[1]}{i}'] = name

            for i, name in enumerate(df_linebible['km.3'], start=4):
                worksheet[f'{col[2]}{i}'] = name

            for i, name in enumerate(df_linebible['km.4'], start=4):
                worksheet[f'{col[3]}{i}'] = name

            for i, name in enumerate(df_linebible['km.5'], start=4):
                worksheet[f'{col[4]}{i}'] = name

            for i, name in enumerate(df_linebible['km.6'], start=4):
                worksheet[f'{col[5]}{i}'] = name
            
            #print("Write complete...")

        except:
            print('Something wrong with {} - check if page is formatted as other pages.'.format(sheet_name))

    # Save changes to workbook
    print("Saving workbook...")
    workbook.save(filename=file_path)
    print("Save complete...")

#VARS
hight_line_range = [(920000, 929999),
                  (930000, 939999)]

freestyle_line_range = [(180000, 190000),
                        (280000, 290000), 
                        (380000, 390000),
                        (480000, 490000),
                        (580000, 590000),
                        (680000, 690000),
                        (780000, 790000),
                        (880000, 890000),
                        (980000, 990000)]

if __name__ == "__main__":
    # VARIABLES
    file_path, path_to_txt      = choose_linebible()
    cwd                         = os.path.dirname(file_path)
    print(cwd)
    backup_linebible(file_path)

    # Load existing workbook
    print("Loading (Linebible) excel workbook...")
    workbook = load_workbook(filename=file_path)
    print("Excel workbook loading complete...")
    
    main()

    print()
    input("Press enter to end program...")